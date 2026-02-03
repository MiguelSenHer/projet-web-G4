from openpyxl import Workbook
from openpyxl.styles import Font
from django.core.files.base import ContentFile
from io import BytesIO


def generate_assembly_xlsx(assembly):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assembly"
    bold = Font(bold=True)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws["A1"] = "Assembly settings"
    ws["A1"].font = bold
    ws["A2"] = "Restriction enzyme"
    ws["B2"] = assembly.restriction_enzyme or ""
    ws["A3"] = "Name"
    ws["B3"] = assembly.name
    ws["A4"] = "Output separator"
    ws["B4"] = assembly.separator or ""
    ws["A9"] = "Assembly composition"
    ws["A9"].font = bold
    ws["B9"] = "Part name ->"
    ws["B10"] = "Part types ->"
    ws["B11"] = "Is optional part ->"
    ws["B12"] = "Part name should be in output name ->"
    ws["B13"] = "Part separator ->"
    input_parts = assembly.inputparts_set.prefetch_related("allowed_types")
    start_col = 3 
    for idx, part in enumerate(input_parts):
        col = start_col + idx
        col_letter = ws.cell(row=9, column=col).column_letter
        ws[f"{col_letter}9"] = part.part_name
        ws[f"{col_letter}10"] = format_types(part.allowed_types.all())
        ws[f"{col_letter}11"] = not part.mandatory
        ws[f"{col_letter}12"] = part.include_in_output_name
        ws[f"{col_letter}13"] = part.separator or ""
    ws["A14"] = "Output plasmid id ↓"
    ws["B14"] = "OutputType (optional) ↓"
    for idx in range(len(input_parts)):
        ws.cell(row=14, column=start_col + idx).value = "↓"
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return ContentFile(buffer.read())

def format_types(types_qs):
    temp = []
    for t in types_qs:
        name = t.type_name
        if len(name) == 1:
            temp.append(name)
        else:
            temp.append(f"[{name}]")       
    return ", ".join(temp)