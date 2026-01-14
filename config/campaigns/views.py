from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.http import FileResponse, Http404
from pathlib import Path
import openpyxl
import zipfile
import uuid
from datetime import datetime
from .models import Assembly
import subprocess


# ============================================================
# Simulator home
# ============================================================
def simulator_home(request):
    options = [
        {
            "label": "LOAD YOUR PLASMID ASSEMBLY TEMPLATE",
            "url": "/campaigns/simulator/upload/",
        },
        {
            "label": "BROWSE PLASMID ASSEMBLY TEMPLATE",
            "url": "/campaigns/simulator/browse/",
        },
    ]
    return render(
        request,
        "campaigns/template.html",
        {
            "options": options,
            "active_page": "simulator",
        },
    )


# ============================================================
# Helpers for XLSX parsing
# ============================================================
def _find_value_right(ws, key, max_rows=80, max_cols=12):
    """
    Cherche une cellule égale à key (case-insensitive, stripped)
    et renvoie la valeur de la cellule à droite (même ligne, col+1).
    """
    key_norm = key.strip().lower()

    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == key_norm:
                return ws.cell(r, c + 1).value

    return None


def _find_part_names(ws, max_rows=120, max_cols=40):
    """
    Trouve la ligne qui commence par "Part name ->"
    et retourne toutes les valeurs à droite (parts).
    """
    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == "part name ->":
                parts = []
                for cc in range(c + 1, max_cols + 1):
                    pv = ws.cell(r, cc).value
                    if pv is None or (isinstance(pv, str) and not pv.strip()):
                        continue
                    parts.append(str(pv).strip())
                return parts

    return []


def _find_cell(ws, label, max_rows=200, max_cols=80):
    target = label.strip().lower()
    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == target:
                return r, c
    return None, None


def _read_row_right(ws, row, start_col, max_cols=200):
    out = []
    for c in range(start_col, max_cols + 1):
        v = ws.cell(row, c).value
        if v is None or (isinstance(v, str) and not v.strip()):
            break
        out.append(str(v).strip())
    return out


def _read_col_down(ws, start_row, col, max_rows=500):
    out = []
    for r in range(start_row, max_rows + 1):
        v = ws.cell(r, col).value
        if v is None or (isinstance(v, str) and not v.strip()):
            break
        out.append(str(v).strip())
    return out


def _cell_str(v):
    if v is None:
        return ""
    return str(v).strip()


def parse_output_plasmids(xlsx_path):
    """
    TEMPLATE Campaign_display_L1.xlsx

    - pIDs: column under "Output plasmid id ↓"
    - types: column under "OutputType (optional) ↓"
    - parts: row right of "Part name ->"
    - values: grid intersection (row pID × column part)

    Returns:
      parts: list[str]
      rows: list[dict] with keys:
        pid, ptype, part_values
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    pid_r, pid_c = _find_cell(ws, "Output plasmid id ↓")
    typ_r, typ_c = _find_cell(ws, "OutputType (optional) ↓")
    part_r, part_c = _find_cell(ws, "Part name ->")

    if not pid_r or not part_r:
        raise ValueError("Missing required headers in template.")

    # Parts headers (row)
    parts = _read_row_right(ws, part_r, part_c + 1)
    if not parts:
        raise ValueError("No parts detected.")

    # pIDs + types (columns)
    pids = _read_col_down(ws, pid_r + 1, pid_c)
    ptypes = (
        _read_col_down(ws, typ_r + 1, typ_c)
        if typ_r else []
    )

    if len(ptypes) < len(pids):
        ptypes += [""] * (len(pids) - len(ptypes))

    rows = []
    for i, pid in enumerate(pids):
        excel_row = pid_r + 1 + i

        part_values = []
        for j in range(len(parts)):
            excel_col = part_c + 1 + j
            part_values.append(
                _cell_str(ws.cell(excel_row, excel_col).value)
            )

        rows.append({
            "pid": pid,
            "ptype": ptypes[i],
            "part_values": part_values,
        })

    return parts, rows


# ============================================================
# Upload template (xlsx)
# ============================================================
def upload_template(request):
    error = None

    # CLEAR TEMPLATE
    if request.method == "POST" and request.POST.get("clear_template") == "1":
        path = request.session.get("uploaded_template_path")
        if path:
            p = Path(path)
            if p.exists():
                p.unlink()

        request.session.pop("uploaded_template_path", None)
        request.session.pop("uploaded_template_name", None)
        request.session.pop("template_is_valid", None)
        request.session.pop("assembly_preview", None)

        return redirect("/campaigns/simulator/upload/")

    # UPLOAD TEMPLATE
    if request.method == "POST" and request.FILES.get("template_file"):
        f = request.FILES["template_file"]

        if not f.name.lower().endswith(".xlsx"):
            error = "Only .xlsx files are allowed."
        else:
            out_dir = Path(settings.MEDIA_ROOT) / "simulator" / "templates"
            out_dir.mkdir(parents=True, exist_ok=True)

            dest = out_dir / f.name
            with open(dest, "wb") as w:
                for chunk in f.chunks():
                    w.write(chunk)

            request.session["uploaded_template_path"] = str(dest)
            request.session["uploaded_template_name"] = f.name

            return redirect("/campaigns/simulator/upload/")

    return render(request, "campaigns/upload.html", {"error": error})


# ============================================================
# Upload template preview / validation
# ============================================================
def upload_template_next(request):
    path_str = request.session.get("uploaded_template_path")
    filename = request.session.get("uploaded_template_name")

    if not path_str:
        return redirect("/campaigns/simulator/upload/")

    try:
        wb = openpyxl.load_workbook(path_str, data_only=True)
        ws = wb.active

        enzyme = _find_value_right(ws, "Restriction enzyme")
        name = _find_value_right(ws, "Name")
        sep = _find_value_right(ws, "Output separator")
        parts = _find_part_names(ws)

        assembly = {
            "name": (
                str(name).strip()
                if name is not None
                else (filename or "Unnamed")
            ),
            "separator": (
                str(sep).strip() if sep is not None else ""
            ),
            "restriction_enzyme": (
                str(enzyme).strip() if enzyme is not None else ""
            ),
            "input_parts": parts,
        }

        # Validation
        missing = []
        if not assembly["restriction_enzyme"]:
            missing.append("Restriction enzyme")
        if not assembly["name"]:
            missing.append("Name")
        if not assembly["separator"]:
            missing.append("Output separator")
        if not assembly["input_parts"]:
            missing.append("Part name -> row")

        error = None
        if missing:
            error = "Template incomplete: missing " + ", ".join(missing)

        is_valid = (len(missing) == 0)
        request.session["template_is_valid"] = is_valid
        request.session["assembly_preview"] = assembly

        return render(
            request,
            "campaigns/upload_preview.html",
            {
                "assembly": assembly,
                "error": error,
                "active_page": "simulator",
            },
        )

    except Exception as e:
        return render(
            request,
            "campaigns/upload_preview.html",
            {
                "assembly": None,
                "error": str(e),
                "active_page": "simulator",
            },
        )


# ============================================================
# Helpers
# ============================================================
def _zip_only_contains_extensions(
    zip_path, allowed_exts, *, allow_empty=False
):
    """
    Check that a ZIP contains ONLY files with extensions in allowed_exts.
    - Ignores directories (entries ending with '/')
    - allowed_exts: set/tuple like {".gb", ".gbk"}
    - allow_empty: if False, empty zip is rejected
    Returns: (ok: bool, error_msg: str|None)
    """
    allowed_exts = {e.lower() for e in allowed_exts}

    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            names = zf.namelist()
    except zipfile.BadZipFile:
        return False, "Uploaded file is not a valid .zip archive (corrupted)."

    file_names = [n for n in names if n and not n.endswith("/")]
    if not file_names and not allow_empty:
        return False, "The .zip archive is empty."

    for n in file_names:
        lower = n.lower()
        if not any(lower.endswith(ext) for ext in allowed_exts):
            ext_list = ', '.join(sorted(allowed_exts))
            msg = f"Invalid file in archive: '{n}'. Only {ext_list} allowed."
            return (False, msg)

    return True, None


def _validate_genbank_zip(zip_path):
    # Extensions GenBank courantes
    allowed = {".gb", ".gbk", ".genbank"}
    return _zip_only_contains_extensions(zip_path, allowed)


def _validate_mapping_zip(zip_path):
    allowed = {".csv", ".tsv", ".txt"}
    return _zip_only_contains_extensions(zip_path, allowed)


# ============================================================
# Simulator inputs (GenBank + mapping)
# ============================================================
def simulator_inputs(request):
    if not request.session.get("template_is_valid", False):
        return redirect("/campaigns/simulator/upload/next/")

    assembly = request.session.get("assembly_preview")
    error = None

    # CLEAR GENBANK
    if request.method == "POST" and request.POST.get("clear_genbank") == "1":
        path = request.session.get("genbank_path")
        if path:
            p = Path(path)
            if p.exists():
                p.unlink()

        request.session.pop("genbank_path", None)
        request.session.pop("genbank_name", None)
        request.session.pop("ok_genbank", None)

        return redirect("/campaigns/simulator/inputs/")

    # CLEAR MAPPING
    if request.method == "POST" and request.POST.get("clear_mapping") == "1":
        path = request.session.get("mapping_path")
        if path:
            p = Path(path)
            if p.exists():
                p.unlink()

        request.session.pop("mapping_path", None)
        request.session.pop("mapping_name", None)
        request.session.pop("ok_mapping", None)

        return redirect("/campaigns/simulator/inputs/")

    # UPLOAD FILES
    if request.method == "POST":
        out_dir = Path(settings.MEDIA_ROOT) / "simulator" / "inputs"
        out_dir.mkdir(parents=True, exist_ok=True)

        # ---------------------------
        # GenBank ZIP (REQUIRED)
        # ---------------------------
        if request.FILES.get("genbank_zip"):
            f = request.FILES["genbank_zip"]

            if not f.name.lower().endswith(".zip"):
                error = "GenBank input must be a .zip archive."
            else:
                dest = out_dir / f.name
                with open(dest, "wb") as w:
                    for chunk in f.chunks():
                        w.write(chunk)

                ok, msg = _validate_genbank_zip(dest)
                if not ok:
                    if dest.exists():
                        dest.unlink()

                    request.session.pop("genbank_path", None)
                    request.session.pop("genbank_name", None)
                    request.session.pop("ok_genbank", None)

                    error = msg
                else:
                    request.session["genbank_path"] = str(dest)
                    request.session["genbank_name"] = f.name
                    request.session["ok_genbank"] = True

        # ---------------------------
        # Mapping file (OPTIONAL)
        # ---------------------------
        if request.FILES.get("mapping_file"):
            f = request.FILES["mapping_file"]
            allowed = (".csv", ".tsv", ".txt", ".zip")

            if not f.name.lower().endswith(allowed):
                error = "Mapping file must be .csv, .tsv, .txt, or .zip."
            else:
                dest = out_dir / f.name
                with open(dest, "wb") as w:
                    for chunk in f.chunks():
                        w.write(chunk)

                if f.name.lower().endswith(".zip"):
                    ok, msg = _validate_mapping_zip(dest)
                    if not ok:
                        if dest.exists():
                            dest.unlink()

                        request.session.pop("mapping_path", None)
                        request.session.pop("mapping_name", None)
                        request.session.pop("ok_mapping", None)

                        error = msg
                    else:
                        request.session["mapping_path"] = str(dest)
                        request.session["mapping_name"] = f.name
                        request.session["ok_mapping"] = True
                else:
                    # .csv/.tsv/.txt -> accepté
                    request.session["mapping_path"] = str(dest)
                    request.session["mapping_name"] = f.name
                    request.session["ok_mapping"] = True

    return render(
        request,
        "campaigns/inputs.html",
        {
            "assembly": assembly,
            "error": error,
            "ok_genbank": request.session.get("ok_genbank"),
            "ok_mapping": request.session.get("ok_mapping"),
            "genbank_name": request.session.get("genbank_name"),
            "mapping_name": request.session.get("mapping_name"),
        },
    )


# ============================================================
# Helper for simulation preview
# ============================================================
def _list_received_files(path_str):
    """
    Return list of received filenames.
    - If path is a zip: return zip content filenames (filtered: no directories)
    - Else: return [basename]
    """
    if not path_str:
        return []

    p = Path(path_str)
    if not p.exists():
        return []

    # ZIP case
    if p.suffix.lower() == ".zip" and zipfile.is_zipfile(p):
        with zipfile.ZipFile(p, "r") as z:
            names = []
            for n in z.namelist():
                # skip directories
                if n.endswith("/"):
                    continue
                names.append(n.split("/")[-1])  # keep basename
            # remove empty names and sort for nicer display
            names = [x for x in names if x]
            return sorted(names)

    # non-zip
    return [p.name]


# ============================================================
# Simulation preview
# ============================================================
def simulation_preview(request):
    if not request.session.get("template_is_valid", False):
        return redirect("/campaigns/simulator/upload/next/")

    genbank_path = request.session.get("genbank_path")
    mapping_path = request.session.get("mapping_path")  # optional

    genbank_files = _list_received_files(genbank_path)
    mapping_files = _list_received_files(mapping_path)

    # Output plasmids table
    template_path = request.session.get("uploaded_template_path")
    output_parts = []
    output_rows = []
    output_error = None

    if template_path:
        try:
            output_parts, output_rows = parse_output_plasmids(template_path)
        except Exception as e:
            output_error = str(e)

    return render(
        request,
        "campaigns/simulation_preview.html",
        {
            "assembly": request.session.get("assembly_preview"),
            "genbank_files": genbank_files,
            "mapping_files": mapping_files,
            "genbank_count": len(genbank_files),
            "mapping_count": len(mapping_files),
            "output_parts": output_parts,
            "output_rows": output_rows,
            "output_error": output_error,
        },
    )


# ============================================================
# Browse templates
# ============================================================
def browse_templates(request):
    assemblies = Assembly.objects.prefetch_related("inputparts_set")
    return render(
        request,
        "campaigns/browse.html",
        {
            "assemblies": assemblies,
            "active_page": "browse",
        },
    )


# ============================================================
# Assembly download
# ============================================================
def assembly_download(request, pk):
    assembly = get_object_or_404(Assembly, pk=pk)
    if not assembly.file_name:
        raise Http404("No file associated with this assembly")
    file_path = (
        Path(settings.BASE_DIR) / "assemblies_files" / assembly.file_name
    )
    print(file_path)
    print(file_path.exists())
    if not file_path.exists():
        raise Http404("File not found")
    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename=assembly.file_name
    )


# ============================================================
# Assembly detail
# ============================================================
def assembly_detail(request, pk):
    assembly = get_object_or_404(Assembly, id=pk)
    input_parts = assembly.inputparts_set.prefetch_related("allowed_types")

    return render(
        request,
        "campaigns/assembly_details.html",
        {
            "assembly": assembly,
            "input_parts": input_parts,
            "active_page": "browse",
        },
    )


# ============================================================
# Helpers to collect zips
# ============================================================
def _extract_zip(zip_path: Path, dest_dir: Path):
    dest_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(dest_dir)


def _collect_files(root: Path):
    return [p for p in root.rglob("*") if p.is_file()]


def _make_zip_from_dir(src_dir: Path, zip_path: Path):
    with zipfile.ZipFile(
        zip_path, "w", compression=zipfile.ZIP_DEFLATED
    ) as zf:
        for p in Path(src_dir).rglob("*"):
            if p.is_file():
                zf.write(p, p.relative_to(src_dir).as_posix())


# ============================================================
# SIMULATION RUN
# ============================================================
def simulation_run(request):
    # Prérequis : template validé + genbank présent
    if not request.session.get("template_is_valid", False):
        return redirect("/campaigns/simulator/upload/next/")

    genbank_zip_path = request.session.get("genbank_path")
    if not genbank_zip_path:
        return redirect("/campaigns/simulator/inputs/")

    template_path = request.session.get("uploaded_template_path")
    if not template_path:
        return redirect("/campaigns/simulator/upload/")

    mapping_path = request.session.get("mapping_path")

    # Run id unique
    run_id = uuid.uuid4().hex[:10]

    # Dossier run
    runs_root = Path(settings.MEDIA_ROOT) / "simulator" / "runs"
    run_dir = runs_root / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    # Dossiers run
    out_dir = run_dir / "output"
    out_dir.mkdir(parents=True, exist_ok=True)

    repo_dir = run_dir / "plasmid_repository" 
    map_dir = run_dir / "mapping"

    # Template
    tpl_path = Path(template_path)

    # Enzyme depuis la session preview (REQUIRED)
    assembly = request.session.get("assembly_preview") or {}
    enzyme = assembly.get("restriction_enzyme")

    status = "ok"
    error = None
    stderr_text = ""
    outputs_zip_url = None

    try:
        # ----------------------
        # GenBank (ZIP required)
        # ----------------------
        _extract_zip(Path(genbank_zip_path), repo_dir)
        gb_files = _collect_files(repo_dir)
        if not gb_files:
            raise ValueError("No GenBank files found after extraction.")

        # ----------------------
        # Mapping (optional zip or single file)
        # ----------------------
        input_parts_files = []
        if mapping_path:
            mp = Path(mapping_path)
            if mp.suffix.lower() == ".zip":
                _extract_zip(mp, map_dir)
                input_parts_files = _collect_files(map_dir)
            else:
                input_parts_files = [mp]

        # ----------------------
        # Build CLI command
        # ----------------------
        cmd = [
            "insillyclo",
            "simulate",
            "--input-template-filled", str(tpl_path),
            "--plasmid-repository", str(repo_dir),
            "--recursive-plasmid-repository",
            "--default-mass-concentration", "200",
            "--restriction-enzyme-gel", enzyme,
            "--output-dir", str(out_dir),
        ]

        # mapping files : repeat option
        for f in input_parts_files:
            cmd += ["--input-parts-file", str(f)]

        # ----------------------
        # Run subprocess
        # ----------------------
        proc = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
        )

        raw = proc.stderr or ""
        stderr_text = raw.split("Traceback", 1)[0].strip()

        if proc.returncode != 0:
            status = "error"
            error = stderr_text
        else:
            zip_path = run_dir / "outputs.zip"
            _make_zip_from_dir(out_dir, zip_path)
            outputs_zip_url = f"/campaigns/simulator/run/{run_id}/download/"

            runs = request.session.get("successful_runs", []) or []
            if run_id not in runs:
                runs.append(run_id)
            request.session["successful_runs"] = runs
            request.session.modified = True

    except Exception as e:
        status = "error"
        error = str(e)

    return render(
        request,
        "campaigns/simulation_run.html",
        {
            "assembly": request.session.get("assembly_preview"),
            "status": status,
            "error": error,
            "stderr_text": stderr_text,
            "run_id": run_id,
            "outputs_zip_url": outputs_zip_url,
        },
    )


def simulation_run_download(request, run_id):
    run_dir = Path(settings.MEDIA_ROOT) / "simulator" / "runs" / run_id
    zip_path = run_dir / "outputs.zip"
    if not zip_path.exists():
        raise Http404("Outputs zip not found.")
    return FileResponse(
        open(zip_path, "rb"),
        as_attachment=True,
        filename=f"outputs_{run_id}.zip"
    )


def simulations_list(request):
    run_ids = request.session.get("successful_runs", []) or []
    runs_root = Path(settings.MEDIA_ROOT) / "simulator" / "runs"

    items = []
    for run_id in run_ids:
        zip_path = runs_root / run_id / "outputs.zip"

        if not zip_path.exists():
            continue

        mtime = zip_path.stat().st_mtime
        items.append({
            "run_id": run_id,
            "status": "SUCCESS",
            "updated_at": datetime.fromtimestamp(mtime),
            "download_url": f"/campaigns/simulator/run/{run_id}/download/",
            "back_url": "/campaigns/simulator/preview/",
        })

    items.sort(key=lambda x: x["updated_at"], reverse=True)

    return render(
        request,
        "campaigns/simulations_list.html",
        {
            "items": items,
            "active_page": "simulations",
        },
    )
